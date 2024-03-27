import openpyxl
import TestData
workbook=openpyxl.load_workbook(r'C:\Users\ACHOPRA\PycharmProjects\FrameWork\TestData\data_driven_test_file.xlsx')
sheet=workbook.active

def getRowCount (file, sheetName):
    workbook=openpyxl.load_workbook (r'C:\Users\ACHOPRA\PycharmProjects\FrameWork\TestData\data_driven_test_file.xlsx')
    sheet=workbook.active
    sheetName=sheet
    return(sheet.max_row)

def getColumnCount(file,sheetName):
    workbook=openpyxl.load_workbook(r'C:\Users\ACHOPRA\PycharmProjects\FrameWork\TestData\data_driven_test_file.xlsx')
    sheet=workbook.active
    sheetName = sheet
    return(sheet.max_column)

def readData(file,sheetName, rownum, columnno):
    workbook=openpyxl.load_workbook(r'C:\Users\ACHOPRA\PycharmProjects\FrameWork\TestData\data_driven_test_file.xlsx')
    sheet=workbook.active
    sheetName = sheet
    return sheet.cell(row=rownum, column=columnno).value

def writeData(file,sheetName, rownum, columnno, data):
    workbook=openpyxl.load_workbook(r'C:\Users\ACHOPRA\PycharmProjects\FrameWork\TestData\data_driven_test_file.xlsx')
    sheet=workbook.active
    sheetName = sheet
    sheet.cell(row=rownum, column=columnno).value = data
    workbook.save(r'C:\Users\ACHOPRA\PycharmProjects\FrameWork\TestData\data_driven_test_file.xlsx')









# cell=sheet.cell(row=3,column=2)
# print(cell.value)
#
# write=sheet.cell(row=7,column=6).value="chopra"
# print(write)
#
# print(sheet.max_row)
# print(sheet.max_column)
#
# print(sheet['C5'].value)